import argparse
import re
import sys

import yaml
from openpyxl import load_workbook

STRING_TEMPLATE_REGEX = re.compile(r"{([^{}]*?)}")

STATEMENT_NAME = "statement"
SUBSTATEMENTS = "substatements"
PLHLD_COLUMN_MAPPING = "placeholder_column_mapping"  # PLACEHOLDER_COLUMN_MAPPING
PLACEHOLDER_ROW_START = "placeholder_row_start"
SHEET_NAME = "sheet_name"

STMT_PLHLDS = "__placeholders__"  # STATEMENT_PLACEHOLDERS


def gen_statement_mapping(config, curr_statement, curr_statement_mapping):
    for match in STRING_TEMPLATE_REGEX.finditer(config[SUBSTATEMENTS][curr_statement]):
        if match.group(1) in config[PLHLD_COLUMN_MAPPING]:
            curr_statement_mapping.setdefault(STMT_PLHLDS, list()).append(match.group(1))
        elif match.group(1) in config[SUBSTATEMENTS]:
            gen_statement_mapping(config, match.group(1), curr_statement_mapping.setdefault(match.group(1), dict()))
        else:
            print(
                f"This key could neither be found in {SUBSTATEMENTS} nor in {PLHLD_COLUMN_MAPPING}: {match.group(1)}",
                file=sys.stderr,
            )
            sys.exit(1)


def gen_substring_value(row, placeholder_mapping, curr_value_mapping, curr_statement_mapping):
    for substatement in curr_statement_mapping:
        if substatement == STMT_PLHLDS:
            for placeholder in curr_statement_mapping[STMT_PLHLDS]:
                curr_value_mapping[placeholder] = row[placeholder_mapping[placeholder]].value
        else:
            gen_substring_value(
                row,
                placeholder_mapping,
                curr_value_mapping.setdefault(substatement, dict()),
                curr_statement_mapping[substatement],
            )


def gen_substring(config, curr_statement_mapping, curr_value_mapping):
    ret_val = {}
    for substatement in curr_statement_mapping:
        if substatement == STMT_PLHLDS:
            ret_val.update({k: curr_value_mapping[k] for k in curr_statement_mapping[STMT_PLHLDS]})
        else:
            ret_val.update(
                {
                    substatement: config[SUBSTATEMENTS][substatement].format(
                        **gen_substring(config, curr_statement_mapping[substatement], curr_value_mapping[substatement])
                    )
                }
            )
    return ret_val


def main():
    parser = argparse.ArgumentParser(
        description="A tool to extract data from excel files and write to complex via yaml config defined structures."
    )
    parser.add_argument("yaml_config")
    parser.add_argument("excel_file")
    parser.add_argument("out_file")

    args = parser.parse_args()

    config = yaml.safe_load(open(args.yaml_config))

    statement_mapping = {}

    # config should never be saved afterwards!
    config[SUBSTATEMENTS][STATEMENT_NAME] = config[STATEMENT_NAME]
    for column in config[PLHLD_COLUMN_MAPPING]:
        config[PLHLD_COLUMN_MAPPING][column] = ord(config[PLHLD_COLUMN_MAPPING][column]) - 65

    gen_statement_mapping(config, STATEMENT_NAME, statement_mapping)

    wb = load_workbook(filename=args.excel_file, read_only=True)

    sheet = wb[config[SHEET_NAME]]

    substring_values = []

    for row in sheet.iter_rows(min_row=config[PLACEHOLDER_ROW_START]):
        value_mapping = {}
        gen_substring_value(row, config[PLHLD_COLUMN_MAPPING], value_mapping, statement_mapping)
        substring_values.append(value_mapping)

    keys = []
    values = {}
    for value_mapping in substring_values:
        keys_set = bool(keys)
        substring = gen_substring(config, statement_mapping, value_mapping)
        for statement in substring:
            values.setdefault(statement, list()).append(substring[statement])
            if not keys_set:
                keys.append(statement)

    with open(args.out_file, "w") as file:
        file.write(config[STATEMENT_NAME].format(**{key: "".join(strings) for key, strings in values.items()}))


if __name__ == "__main__":
    main()
